"""
Module pour générer les fichiers Excel de sortie avec table matrice
"""
import pandas as pd
from pathlib import Path
from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class MatrixGenerator:
    """Génère des fichiers Excel avec mise en forme de table matrice"""
    
    def __init__(self, output_path: str):
        """
        Initialise le générateur
        
        Args:
            output_path: Chemin du fichier de sortie
        """
        self.output_path = Path(output_path)
        self.wb = Workbook()
        self.ws = self.wb.active
        
    def create_from_dataframe(
        self,
        df: pd.DataFrame,
        sheet_name: str = "Catalogue",
        apply_styling: bool = True
    ):
        """
        Crée une table matrice depuis un DataFrame
        
        Args:
            df: DataFrame avec les données
            sheet_name: Nom de la feuille
            apply_styling: Appliquer le style de table
        """
        self.ws.title = sheet_name
        
        # Écrit les données
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = self.ws.cell(row=r_idx, column=c_idx, value=value)
                
                if apply_styling:
                    # Style pour les en-têtes
                    if r_idx == 1:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(vertical="top")
                    
                    # Bordures
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    cell.border = thin_border
        
        # Auto-ajuste les colonnes
        if apply_styling:
            self._auto_adjust_columns()
    
    def _auto_adjust_columns(self):
        """Ajuste automatiquement la largeur des colonnes"""
        for column in self.ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Max 50
            self.ws.column_dimensions[column_letter].width = adjusted_width
    
    def add_summary_sheet(self, summary_data: Dict[str, any]):
        """
        Ajoute une feuille de résumé
        
        Args:
            summary_data: Dictionnaire avec les infos de résumé
        """
        summary_ws = self.wb.create_sheet("Résumé", 0)
        
        # Titre
        summary_ws['A1'] = "Résumé du traitement"
        summary_ws['A1'].font = Font(size=14, bold=True)
        
        row = 3
        for key, value in summary_data.items():
            summary_ws[f'A{row}'] = key
            summary_ws[f'B{row}'] = value
            summary_ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        summary_ws.column_dimensions['A'].width = 30
        summary_ws.column_dimensions['B'].width = 40
    
    def add_mapping_sheet(self, column_mapping: Dict[str, Dict[str, any]]):
        """
        Ajoute une feuille avec le mapping des colonnes
        
        Args:
            column_mapping: Mapping avec colonnes et scores de confiance
        """
        mapping_ws = self.wb.create_sheet("Mapping colonnes")
        
        # En-têtes
        headers = ["Colonne cible", "Colonne source", "Confiance"]
        for c_idx, header in enumerate(headers, 1):
            cell = mapping_ws.cell(row=1, column=c_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Données
        row = 2
        for target_col, mapping_info in column_mapping.items():
            mapping_ws.cell(row=row, column=1, value=target_col)
            
            if isinstance(mapping_info, dict):
                mapping_ws.cell(row=row, column=2, value=mapping_info.get('column', 'N/A'))
                confidence = mapping_info.get('confidence', 0)
                mapping_ws.cell(row=row, column=3, value=f"{confidence:.0%}")
            else:
                mapping_ws.cell(row=row, column=2, value=mapping_info or 'N/A')
                mapping_ws.cell(row=row, column=3, value="N/A")
            
            row += 1
        
        # Ajuste les colonnes
        for col in ['A', 'B', 'C']:
            mapping_ws.column_dimensions[col].width = 25
    
    def save(self):
        """Sauvegarde le fichier Excel"""
        try:
            self.output_path.parent.mkdir(parents=True, exist_ok=True)
            self.wb.save(self.output_path)
            print(f"✓ Fichier généré: {self.output_path}")
            return True
        except Exception as e:
            print(f"✗ Erreur lors de la sauvegarde: {e}")
            return False
    
    @staticmethod
    def create_matrix_excel(
        df: pd.DataFrame,
        output_path: str,
        column_mapping: Optional[Dict] = None,
        summary_data: Optional[Dict] = None
    ):
        """
        Méthode helper pour créer un Excel complet en une seule fois
        
        Args:
            df: DataFrame avec les données
            output_path: Chemin de sortie
            column_mapping: Mapping des colonnes (optionnel)
            summary_data: Données de résumé (optionnel)
        """
        generator = MatrixGenerator(output_path)
        generator.create_from_dataframe(df, apply_styling=True)
        
        if summary_data:
            generator.add_summary_sheet(summary_data)
        
        if column_mapping:
            generator.add_mapping_sheet(column_mapping)
        
        return generator.save()
