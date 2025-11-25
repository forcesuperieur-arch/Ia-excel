"""
Module pour injecter les données dans un template Excel existant
"""
import pandas as pd
from pathlib import Path
from typing import Dict, Optional, List, Tuple
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


class TemplateInjector:
    """Injecte des données dans un template Excel en préservant la mise en forme"""
    
    def __init__(self, template_path: str):
        """
        Initialise l'injecteur avec un template
        
        Args:
            template_path: Chemin vers le fichier Excel template
        """
        self.template_path = Path(template_path)
        self.wb = load_workbook(self.template_path)
        self.ws = None
        
    def select_sheet(self, sheet_name: Optional[str] = None, sheet_index: int = 0):
        """
        Sélectionne la feuille où injecter les données
        
        Args:
            sheet_name: Nom de la feuille (ou None pour utiliser l'index)
            sheet_index: Index de la feuille si sheet_name est None
        """
        if sheet_name:
            self.ws = self.wb[sheet_name]
        else:
            self.ws = self.wb.worksheets[sheet_index]
        
        print(f"✓ Feuille sélectionnée: {self.ws.title}")
    
    def find_injection_zone(self, marker: str = "{{DATA}}") -> Optional[Tuple[int, int]]:
        """
        Cherche un marqueur dans le template pour savoir où injecter
        
        Args:
            marker: Marqueur à rechercher (ex: "{{DATA}}", "{{CATALOGUE}}")
            
        Returns:
            Tuple (row, col) de la position du marqueur, ou None
        """
        if not self.ws:
            raise ValueError("Aucune feuille sélectionnée. Utilisez select_sheet() d'abord.")
        
        for row in self.ws.iter_rows():
            for cell in row:
                if cell.value and marker in str(cell.value):
                    print(f"✓ Marqueur '{marker}' trouvé en {cell.coordinate}")
                    return (cell.row, cell.column)
        
        print(f"⚠ Marqueur '{marker}' non trouvé")
        return None
    
    def auto_detect_table_start(self) -> Tuple[int, int]:
        """
        Détecte automatiquement le début d'un tableau (première ligne avec plusieurs en-têtes)
        
        Returns:
            Tuple (row, col) du début du tableau
        """
        if not self.ws:
            raise ValueError("Aucune feuille sélectionnée")
        
        for row_idx, row in enumerate(self.ws.iter_rows(max_row=20), start=1):
            # Compte les cellules non vides dans la ligne
            non_empty = sum(1 for cell in row if cell.value is not None)
            
            # Si on trouve une ligne avec plusieurs cellules (probable en-tête)
            if non_empty >= 3:
                print(f"✓ Début de tableau détecté à la ligne {row_idx}")
                return (row_idx, 1)
        
        print("⚠ Pas de tableau détecté, utilisation de A1")
        return (1, 1)
    
    def inject_data(
        self,
        df: pd.DataFrame,
        start_row: Optional[int] = None,
        start_col: Optional[int] = None,
        include_headers: bool = True,
        clear_existing: bool = False,
        marker: str = "{{DATA}}"
    ):
        """
        Injecte les données du DataFrame dans le template
        
        Args:
            df: DataFrame à injecter
            start_row: Ligne de début (auto-détecté si None)
            start_col: Colonne de début (auto-détecté si None)
            include_headers: Inclure les en-têtes du DataFrame
            clear_existing: Effacer les données existantes
            marker: Marqueur à chercher pour l'injection
        """
        if not self.ws:
            raise ValueError("Aucune feuille sélectionnée")
        
        # Détection de la position d'injection
        if start_row is None or start_col is None:
            # Cherche d'abord un marqueur
            position = self.find_injection_zone(marker)
            
            if position:
                start_row, start_col = position
            else:
                # Auto-détection du tableau
                start_row, start_col = self.auto_detect_table_start()
                
                # Si on a trouvé un en-tête, on injecte juste en dessous
                if start_row > 1 or not include_headers:
                    start_row += 1
        
        # Efface les données existantes si demandé
        if clear_existing:
            self._clear_area(start_row, start_col, df.shape[0] + 10, df.shape[1])
        
        # Injection des données
        rows_data = dataframe_to_rows(df, index=False, header=include_headers)
        
        for r_idx, row in enumerate(rows_data):
            for c_idx, value in enumerate(row):
                cell = self.ws.cell(
                    row=start_row + r_idx,
                    column=start_col + c_idx,
                    value=value
                )
        
        print(f"✓ {len(df)} lignes injectées à partir de {get_column_letter(start_col)}{start_row}")
    
    def inject_with_column_mapping(
        self,
        df: pd.DataFrame,
        column_mapping: Dict[str, str],
        header_row: int = 1
    ):
        """
        Injecte les données en mappant les colonnes du DataFrame aux colonnes du template
        
        Args:
            df: DataFrame source
            column_mapping: Dict {colonne_template: colonne_df}
            header_row: Ligne contenant les en-têtes dans le template
        """
        if not self.ws:
            raise ValueError("Aucune feuille sélectionnée")
        
        # Lit les en-têtes du template
        template_headers = {}
        for cell in self.ws[header_row]:
            if cell.value:
                template_headers[str(cell.value).strip()] = cell.column
        
        print(f"✓ En-têtes du template: {list(template_headers.keys())}")
        
        # Log du mapping et des colonnes disponibles
        print(f"✓ Colonnes disponibles dans le DataFrame: {list(df.columns)}")
        print(f"✓ Mapping à appliquer: {column_mapping}")
        
        # Injection ligne par ligne
        data_start_row = header_row + 1
        
        for df_row_idx, df_row in df.iterrows():
            excel_row = data_start_row + df_row_idx
            
            for template_col, df_col in column_mapping.items():
                if template_col in template_headers:
                    if df_col in df.columns:
                        col_idx = template_headers[template_col]
                        value = df_row[df_col]
                        
                        self.ws.cell(
                            row=excel_row,
                            column=col_idx,
                            value=value
                        )
                    else:
                        if df_row_idx == 0:  # Log seulement pour la première ligne
                            print(f"⚠ Colonne '{df_col}' non trouvée dans le DataFrame")
                else:
                    if df_row_idx == 0:
                        print(f"⚠ Colonne template '{template_col}' non trouvée dans les en-têtes")
        
        print(f"✓ {len(df)} lignes injectées avec mapping de colonnes")
    
    def _clear_area(self, start_row: int, start_col: int, num_rows: int, num_cols: int):
        """Efface une zone du template"""
        for r in range(start_row, start_row + num_rows):
            for c in range(start_col, start_col + num_cols):
                self.ws.cell(row=r, column=c, value=None)
    
    def list_sheets(self) -> List[str]:
        """Liste toutes les feuilles du template"""
        return self.wb.sheetnames
    
    def get_headers_from_row(self, row: int = 1) -> List[str]:
        """Récupère les en-têtes d'une ligne spécifique"""
        if not self.ws:
            raise ValueError("Aucune feuille sélectionnée")
        
        headers = []
        for cell in self.ws[row]:
            if cell.value:
                headers.append(str(cell.value).strip())
        
        return headers
    
    def save(self, output_path: str):
        """
        Sauvegarde le template avec les données injectées
        
        Args:
            output_path: Chemin du fichier de sortie
        """
        try:
            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            self.wb.save(output_path)
            print(f"✓ Fichier sauvegardé: {output_path}")
            return True
        except Exception as e:
            print(f"✗ Erreur lors de la sauvegarde: {e}")
            return False
    
    @staticmethod
    def inject_into_template(
        template_path: str,
        df: pd.DataFrame,
        output_path: str,
        sheet_name: Optional[str] = None,
        marker: str = "{{DATA}}",
        include_headers: bool = True,
        column_mapping: Optional[Dict[str, str]] = None
    ) -> bool:
        """
        Méthode helper pour injection rapide
        
        Args:
            template_path: Chemin du template
            df: DataFrame à injecter
            output_path: Chemin de sortie
            sheet_name: Nom de la feuille cible
            marker: Marqueur d'injection
            include_headers: Inclure les en-têtes
            column_mapping: Mapping optionnel des colonnes
            
        Returns:
            True si succès
        """
        injector = TemplateInjector(template_path)
        injector.select_sheet(sheet_name)
        
        if column_mapping:
            # Injection avec mapping
            injector.inject_with_column_mapping(df, column_mapping, header_row=1)
        else:
            # Injection simple
            injector.inject_data(
                df,
                include_headers=include_headers,
                marker=marker
            )
        
        return injector.save(output_path)
